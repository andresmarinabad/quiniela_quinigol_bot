from telegram import Update, Document
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
import os
import pandas as pd
import requests
import shutil
from openpyxl import load_workbook

EXCEL_DIR = "excel"
os.makedirs(EXCEL_DIR, exist_ok=True)

esperando_archivo = set()

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
FUTBALL_TOKEN = os.getenv("FUTBALL_API_TOKEN")

BASE_URL = 'https://api.football-data.org/v4'
HEADERS = {
    'X-Auth-Token': FUTBALL_TOKEN
}

def obtener_jornada_actual_pd():
    url = f'{BASE_URL}/competitions/PD/matches?status=FINISHED'
    response = requests.get(url, headers=HEADERS)

    if response.status_code != 200:
        print(f'Error {response.status_code}: {response.text}')
        return

    datos = response.json()
    partidos_por_jornada = {}

    for partido in datos['matches']:
        jornada = partido['matchday']
        if jornada not in partidos_por_jornada:
            partidos_por_jornada[jornada] = []
        partidos_por_jornada[jornada].append(partido)

    jornada_actual = max(partidos_por_jornada.keys())
    print(f"Resultados de la jornada {jornada_actual}:\n")

    for partido in partidos_por_jornada[jornada_actual]:
        local = partido['homeTeam']['name']
        visitante = partido['awayTeam']['name']
        resultado = f"{partido['score']['fullTime']['home']} - {partido['score']['fullTime']['away']}"
        print(f"{local} {resultado} {visitante}")


def generar_nueva_plantilla_jornada():
    url = f'{BASE_URL}/competitions/PD/matches?status=SCHEDULED'
    response = requests.get(url, headers=HEADERS)

    if response.status_code != 200:
        print(f'Error {response.status_code}: {response.text}')
        return

    datos = response.json()
    partidos_por_jornada = {}

    for partido in datos['matches']:
        jornada = partido['matchday']
        if jornada not in partidos_por_jornada:
            partidos_por_jornada[jornada] = []
        partidos_por_jornada[jornada].append(partido)

    proxima_jornada = min(partidos_por_jornada.keys())

    #shutil.copyfile('template.xlsx', 'output.xlsx')

    wb = load_workbook('template.xlsx')
    ws = wb.active
    fila = 6

    ws.cell(row=1, column=1, value=f'Jornada {proxima_jornada}')

    for partido in partidos_por_jornada[proxima_jornada]:
        local = partido['homeTeam']['name']
        visitante = partido['awayTeam']['name']

        ws.cell(row=fila, column=1, value=local)
        ws.cell(row=fila, column=2, value=visitante)
        fila += 1

    wb.save('output.xlsx')


async def descargar_plantilla(update: Update, context: ContextTypes.DEFAULT_TYPE):
    generar_nueva_plantilla_jornada()
    await update.message.reply_document(document=open('output.xlsx', "rb"), filename="jornada_quiniela.xlsx")

async def nueva_jornada(update: Update, context: ContextTypes.DEFAULT_TYPE):
    esperando_archivo.add(update.effective_user.id)
    await update.message.reply_text("Envíame el archivo Excel (formato .xlsx o .xls).")

async def resultados(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Este comando mostrara los resultados de la quiniela con puntuaciones")


async def recibir_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    document: Document = update.message.document

    if user_id not in esperando_archivo:
        await update.message.reply_text("Por favor, primero usa el comando /cargar_excel.")
        return

    if not document.file_name.endswith((".xlsx", ".xls")):
        await update.message.reply_text("El archivo debe ser Excel (.xlsx o .xls).")
        return

    file = await context.bot.get_file(document.file_id)
    await file.download_to_drive(custom_path=document.file_name)

    esperando_archivo.remove(user_id)
    await update.message.reply_text("Recibido y guardado :)")

def main():

    if not TOKEN:
        print('no token')
        return

    if not FUTBALL_TOKEN:
        print('no api futbol token')
        return

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("nueva_jornada", nueva_jornada))
    app.add_handler(CommandHandler("resultados", resultados))
    app.add_handler(CommandHandler("descargar_plantilla", descargar_plantilla))
    app.add_handler(MessageHandler(filters.Document.ALL, recibir_documento))

    print("Bot en marcha...")
    app.run_polling()

if __name__ == "__main__":
    main()
